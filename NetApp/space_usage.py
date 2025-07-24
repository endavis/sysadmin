"""
This goes through all volumes and puts the data in 1 of 4 buckets

CVO HA, RW volume
CVO HA, DP volume
CVO, RW volume
CVO, DP volume
"""

import logging
import pprint
import traceback
from datetime import datetime

from netapp_ontap import HostConnection, utils
from netapp_ontap.resources import Disk, Cluster, Volume, Node
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import xlsxwriter


from libs.config import Config
from libs.parseargs import argp
from libs.log import setup_logger
from libs.size_utils import approximate_size_specific, convert_size
from libs.excel import fix_column_widths, set_font, filter_sheet, get_column_mapping

setup_logger()

#logging.basicConfig(level=logging.DEBUG)
#utils.LOG_ALL_API_CALLS = 1
#=SUMIFS(Volumes!$L$2:$L$478,Volumes!$B$2:$B$478,"MRFS")

config = None

APP = None

class AppClass:
    def __init__(self, name, clusters):
        self.name = name
        self.cluster_details = clusters
        self.clusterdata = {}        
        self.wb = Workbook()
        self.usage_sheet = self.wb.active
        self.usage_sheet.title = "Usage"
        self.volume_sheet = self.wb.create_sheet('Volumes')
        self.volume_sheet.append(["Division", "BU", "Cluster", "App", "Environment", "SubApp", "Cloud", "Region", "Volume", "Tags", "State", "Provisioned Size [TiB]", "Used Size [TiB]"])
        self.divisions = {}
        self.build_app()

    def build_app(self):
        for item in self.cluster_details:
            div = self.cluster_details[item]['div']
            bu = self.cluster_details[item]['bu']
            app = self.cluster_details[item]['app']
            env = self.cluster_details[item]['env']
            subapp = self.cluster_details[item]['subapp']
            cloud = self.cluster_details[item]['cloud']
            region = self.cluster_details[item]['region']
            if div not in self.divisions:
                self.divisions[div] = {}
            if bu not in self.divisions[div]:
                self.divisions[div][bu] = {}
            if app not in self.divisions[div][bu]:
                self.divisions[div][bu][app] = {}
            if env not in self.divisions[div][bu][app]:
                self.divisions[div][bu][app][env] = {}
            if subapp not in self.divisions[div][bu][app][env]:
                self.divisions[div][bu][app][env][subapp] = {}
            if cloud not in self.divisions[div][bu][app][env][subapp]:
                self.divisions[div][bu][app][env][subapp][cloud] = []
            if region not in self.divisions[div][bu][app][env][subapp][cloud]:
                self.divisions[div][bu][app][env][subapp][cloud].append(region)

            self.clusterdata[item] = ClusterData(item, self, **self.cluster_details[item])

        pprint.pprint(self.divisions)

    def go(self):
        for cluster in self.clusterdata.values():
            cluster.gather_data()

        self.build_usage_sheet()

        set_font(self.wb)
        filter_sheet(self.volume_sheet)        
        filter_sheet(self.usage_sheet)        
        fix_column_widths(self.wb)

        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")            
        filename = config.output_dir / f"volumes_{timestamp}.xlsx"
        self.wb.save(filename)

        self.add_usage_table()
        logging.info(f"Data saved to {filename}")       

    def add_usage_table(self):
        # Define the table range (e.g., A1:C4)
        # print(f"Tableref: A1:{get_column_letter(self.usage_sheet.max_column)}{self.usage_sheet.max_row}")
        # table = Table(displayName="Usage", ref=f"A1:{get_column_letter(self.usage_sheet.max_column)}{self.usage_sheet.max_row}")

        # # Add a style to the table
        # style = TableStyleInfo(
        #     name="TableStyleMedium9", showFirstColumn=False,
        #                showLastColumn=False, showRowStripes=True, showColumnStripes=False
        # )
        # table.tableStyleInfo = style

        # Add the table to the worksheet
        # self.usage_sheet.add_table(table)
        # column_mapping = get_column_mapping(self.usage_sheet)
        # provisioned_col = column_mapping[provision_col_header]
        # used_col = column_mapping[used_col_header]

        # row_num = self.usage_sheet.max_row + 1
        # sanitized_prov_col_header = provision_col_header.replace('[', "'[").replace(']', "']")
        # sanitized_used_col_header = used_col_header.replace('[', "'[").replace(']', "']")
        # self.usage_sheet[f"{provisioned_col}{row_num}"] = f"=SUBTOTAL(109,[{sanitized_prov_col_header}])"
        # self.usage_sheet[f"{used_col}{row_num}"] = f"=SUBTOTAL(109,[{sanitized_used_col_header}])"        
        pass

    def build_usage_sheet(self):
        provision_col_header = "Provisioned [TiB]"
        used_col_header = "Used [TiB]"
        self.usage_sheet.append(["Division", "BU", "App", "Environment", "SubApp", "Cloud", "Region", provision_col_header, used_col_header])
        last_row = self.volume_sheet.max_row
        """
            div = self.cluster_details[item]['div']
            bu = self.cluster_details[item]['bu']
            app = self.cluster_details[item]['app']
            env = self.cluster_details[item]['env']
            subapp = self.cluster_details[item]['subapp']
            cloud = self.cluster_details[item]['cloud']
            region = self.cluster_details[item]['region']
        """
        for div in self.divisions:
            # self.usage_sheet.append([div, "", "", "", "", "", "",
            #                         f'=SUMIFS(Volumes!$L$2:$L${last_row},Volumes!$A$2:$A${last_row},"{div}")',
            #                         f'=SUMIFS(Volumes!$M$2:$M${last_row},Volumes!$A$2:$A${last_row},"{div}")'])

            for bu in self.divisions[div]:
                # self.usage_sheet.append([div, bu, "", "", "", "", "",
                #                     f'=SUMIFS(Volumes!$L$2:$L${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}")',
                #                     f'=SUMIFS(Volumes!$M$2:$M${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}")'])

                for app in self.divisions[div][bu]:
                    # if app:
                        # self.usage_sheet.append([div, bu, app, "", "", "", "",
                        #                  f'=SUMIFS(Volumes!$L$2:$L${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}")',
                        #                  f'=SUMIFS(Volumes!$M$2:$M${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}")'])

                    for env in self.divisions[div][bu][app]:
                        # self.usage_sheet.append([div, bu, app, env, "", "", "",
                        #                     f'=SUMIFS(Volumes!$L$2:$L${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}")',
                        #                     f'=SUMIFS(Volumes!$M$2:$M${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}")'])

                        for subapp in self.divisions[div][bu][app][env]:
                            # if subapp:
                            #     self.usage_sheet.append([div, bu, app, env, subapp, "", "",
                            #                     f'=SUMIFS(Volumes!$L$2:$L${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}",Volumes!$F$2:$F${last_row},"{subapp}")',
                            #                     f'=SUMIFS(Volumes!$M$2:$M${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}",Volumes!$F$2:$F${last_row},"{subapp}")'])

                            for cloud in self.divisions[div][bu][app][env][subapp]:
                                # self.usage_sheet.append([div, bu, app, env, subapp, cloud, "",
                                #                     f'=SUMIFS(Volumes!$L$2:$L${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}",Volumes!$F$2:$F${last_row},"{subapp}",Volumes!$G$2:$G${last_row},"{cloud}")',
                                #                     f'=SUMIFS(Volumes!$M$2:$M${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}",Volumes!$F$2:$F${last_row},"{subapp}",Volumes!$G$2:$G${last_row},"{cloud}")'])

                                for region in self.divisions[div][bu][app][env][subapp][cloud]:
                                    self.usage_sheet.append([div, bu, app, env, subapp, cloud, region, 0, 0])
                                                        # f'=SUMIFS(Volumes!$L$2:$L${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}",Volumes!$F$2:$F${last_row},"{subapp}",Volumes!$G$2:$G${last_row},"{cloud}",Volumes!$H$2:$H${last_row},"{region}")',
                                                        # f'=SUMIFS(Volumes!$M$2:$M${last_row},Volumes!$A$2:$A${last_row},"{div}",Volumes!$B$2:$B${last_row},"{bu}",Volumes!$D$2:$D${last_row},"{app}",Volumes!$E$2:$E${last_row},"{env}",Volumes!$F$2:$F${last_row},"{subapp}",Volumes!$G$2:$G${last_row},"{cloud}",Volumes!$H$2:$H${last_row},"{region}")'])

class ClusterData:
    def __init__(self, clustername: str, app_instance: AppClass, **kwargs):
        self.name = clustername
        self.cluster_type = ''
        for name, value in kwargs.items():
            setattr(self, name, value)
        self.app_instance = app_instance

    def gather_data(self):
        logging.info(f'Gathering data for {self.name}')
        try:
            with HostConnection(self.ip, username='cvomon', password=config.settings['users']['cvomon']['enc'], verify=False):
                cluster = Cluster()

                cluster.get()
                volume_args = {}
                volume_args['is_svm_root'] = False
                volume_args['fields'] = '*'
                                        
                volumes = list(Volume.get_collection(**volume_args))
                for volume in volumes:
                    if volume['state'] == 'offline':
                        used = 0
                    else:
                        used = float(f"{approximate_size_specific(volume['space']['used'], 'TiB', withsuffix=False):.2f}")
                    self.app_instance.volume_sheet.append([self.div,
                                                 self.bu,
                                                 self.name,
                                                 self.app,
                                                 self.env,
                                                 self.subapp,
                                                 self.cloud,
                                                 self.region,
                                                 volume['name'],
                                                 volume['state'],
                                                 ",".join(self.tags),
                                                 float(f"{approximate_size_specific(volume['size'], 'TiB', withsuffix=False):.2f}"),
                                                 used])
                    
                    # set to numbers format with 2 decimal places
                    self.app_instance.volume_sheet[f'K{self.app_instance.volume_sheet.max_row}'].number_format = '#,##0.00'
                    self.app_instance.volume_sheet[f'L{self.app_instance.volume_sheet.max_row}'].number_format = '#,##0.00'

        except Exception as e:
            logging.error(f"Could not retrieve data for {self.name} {volume['name']} {e}", exc_info=e)


if __name__ == '__main__':
    args = argp(description="gather volume and cluster stats, provisioned size and savings if changing to 80% and 90% autosize thresholds")
    config = Config(args.config_dir)

    #filter = '{"bu":"Professional", "app":"Axcess", "env":"Prod"}'
    items = config.get_clusters(args.filter)    
    #items = config.get_clusters(filter)

    APP = AppClass('Provisioned', items)
    APP.go()

