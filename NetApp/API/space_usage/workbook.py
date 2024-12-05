import logging

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.filters import (
    FilterColumn,
    CustomFilter,
    CustomFilters,
    DateGroupItem,
    Filters,
)

letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

class SpaceWorkbook(Workbook):
    def __init__(self, path, **kwargs):
        super().__init__(**kwargs)
        self.swb_path = path
        self.setup_totals_sheet()
        self.save()
        self.apps = {}
        self.envs = {}
        self.swb_clusters = {}

    def save(self):
        super().save(self.swb_path)

    def close(self):
        for item in self.swb_clusters.values():
            item.close()
        self.update_sheets()
        self.save()

    def setup_totals_sheet(self):
        self.active.title = 'Totals'
        ws = self.get_sheet_by_name('Totals')
        ws.append(['App', 'ENV', 'Region', 'Cluster', 'Type', 'Total Data (GB)', 'Total Hot Data (GB)', 'Total Cold Data (GB)'])

    def addcluster(self, app, environment, clustername):
        self.swb_clusters[clustername] = ClusterSheet(app, environment, clustername, self)
        return self.swb_clusters[clustername]

    def addenvironment(self, app, environment):
        if app not in self.apps:
            self.apps[app] = {}
        if app not in self.sheetnames:
            self.create_sheet(title=app, index=1)
        if environment not in self.apps[app]:
            self.apps[app][environment] = {}
        sheet_name = f'{app}-{environment}'
        if sheet_name not in self.sheetnames:
            self.create_sheet(title=sheet_name, index=2)
        self.save()

    def addcluster(self, app, environment, cluster, ip):
        if cluster not in self.app[app][environment]:
            self.app[app][environment][cluster] = ip

    def update_sheets(self):
        ...
        
    def build_environment_sheets(self):
        ...

    def update_app_sheets(self):
        ...

    def build_totals_sheet(self):
        self.fix_column_widths('Totals')
        self.set_column_width_to_header('Totals', [6, 7, 8])
        ws = self.get_sheet_by_name('Totals by Datatype')
        filters = ws.auto_filter
        filters.ref = f"A1:H{ws.max_row}"        
        ws.append(['', '', '', '', 'Totals', f'=SUBTOTAL(109, F2:F{ws.max_row})',
                       f'=SUBTOTAL(109, G2:G{ws.max_row})',
                       f'=SUBTOTAL(109, H2:H{ws.max_row})'])
        
        row = ws.max_row
        ws[f'C{row}'].number_format = '#,##0.00'
        ws[f'D{row}'].number_format = '#,##0.00'
        ws[f'E{row}'].number_format = '#,##0.00'

    def set_column_width_to_header(self, sheet, columns):
        ws = self.get_sheet_by_name(sheet)
        for column in columns:
            header_text = ws.cell(row=1, column=column).value
            header_text_length = len(header_text)
            if header_text_length < 5:
                text_length = header_text_length + 4
            else:
                text_length = header_text_length + 3
            ws.column_dimensions[letters[column - 1]].width = text_length

    def fix_column_widths(self, sheet_name):
        ws = self.get_sheet_by_name(sheet_name)
        for idx, column_cells in enumerate(ws.columns):
            header_text_width = len(str(column_cells[0].value))
            data_text_width = max(len(str(cell.value)) for cell in column_cells[1:])
            if header_text_width > data_text_width:
                length = header_text_width
                if length < 5:
                    length = length + 4
                else:
                    length = length + 3
            else:
                length = data_text_width
                length = length * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = length
        for idx, col in enumerate(ws.columns, 1):
            ws.column_dimensions[get_column_letter(idx)].auto_size = True
        ColumnDimension(ws, bestFit=True)        


class ClusterSheet:
    def __init__(self, app, environment, name, spacewb):
        self.app = app
        self.environment = environment
        self.name = name
        self.wb = spacewb
        self.region = self.name[1:4]
        self.row = 1
        self.columns = ['App', 'Environment', 'Cluster', 'Region', 'Volume', 'Type', 'Total Used (GB)', 'Hot (GB)', 'Cold (GB)']
        self.sheet = self.wb.create_sheet(title=self.name)
        self.types_seen = []
        self.create_header()

    def create_header(self):
        self.sheet.append(self.columns)
        self.wb.save()

    def addvolume(self, name, vtype, totalsize, hotsize, coldsize):
        logging.debug(f'Cluster {self.name} : volume {name} of type {vtype} = {totalsize=} {hotsize=} {coldsize=}')        
        self.row += 1
        if vtype not in self.types_seen:
            self.types_seen.append(vtype)
        self.sheet.append([self.app, self.environment, self.name, self.region, name, vtype, round(totalsize, 2), round(hotsize, 2), round(coldsize, 2)])        
        self.sheet[f'E{self.row}'].number_format = '#,##0.00'
        self.sheet[f'F{self.row}'].number_format = '#,##0.00'
        self.sheet[f'G{self.row}'].number_format = '#,##0.00'
        self.wb.save()

    def close(self):
        filters = self.sheet.auto_filter
        last_row = self.sheet.max_row
        filters.ref = f"A1:G{self.sheet.max_row}"

        self.sheet.append(['', '', '', 'Totals', f'=SUBTOTAL(109, E2:E{self.sheet.max_row})',
                       f'=SUBTOTAL(109, F2:F{self.sheet.max_row})',
                       f'=SUBTOTAL(109, G2:G{self.sheet.max_row})'])
        
        row = self.sheet.max_row
        self.sheet[f'E{row}'].number_format = '#,##0.00'
        self.sheet[f'F{row}'].number_format = '#,##0.00'
        self.sheet[f'G{row}'].number_format = '#,##0.00'
        
        self.wb.fix_column_widths(self.name) 

        # Update Environment sheet
        sheet_name = f'{self.app}-{self.environment}'
        env_sheet = self.wb.get_sheet_by_name(sheet_name)
        row = env_sheet.max_row + 1
        for item in self.types_seen:
            env_sheet.append([self.name, item, f'=SUMIF({self.name}!C2:C{last_row}, B{row}, {self.name}!D2:D{last_row})', 
                                f'=SUMIF({self.name}!C2:C{last_row}, B{row}, {self.name}!E2:E{last_row})', 
                                f'=SUMIF({self.name}!C2:C{last_row}, B{row}, {self.name}!F2:F{last_row})'])
            env_sheet[f'C{row}'].number_format = '#,##0.00'
            env_sheet[f'D{row}'].number_format = '#,##0.00'
            env_sheet[f'E{row}'].number_format = '#,##0.00'
            row += 1

        # Update App sheet



        # Update Totals Sheet


        self.wb.save()
