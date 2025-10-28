"""

"""
import pprint
import logging
import sys
import pathlib
from datetime import datetime

#from workbook import SpaceWorkbook
from netapp_ontap import HostConnection, utils # pyright: ignore[reportPrivateImportUsage]
from netapp_ontap.resources import ClientLock, Volume
from openpyxl import Workbook

from libs.config import Config
from libs.parseargs import argp
from libs.log import setup_logger


logger = setup_logger()
#utils.LOG_ALL_API_CALLS = 1

script_name = pathlib.Path(__file__).stem

APP = None

class AppClass:
    def __init__(self, name, clusters, config):
        self.config = config
        self.name = name
        self.cluster_details = clusters
        self.clusterdata = {}
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        self.filename = config.output_dir / f"locks_{timestamp}.xlsx"
        self.wb = Workbook()
        self.build_app()

    def build_app(self):
        for item in self.cluster_details:
            self.clusterdata[item] = ClusterData(item, self, **self.cluster_details[item])

    def go(self):
        for cluster in self.clusterdata.values():
            cluster.gather_data()
            cluster.output_data()

        logging.info(f"{self.wb.sheetnames}")
        self.wb.save(self.filename)
        logging.info(f"output saved to {self.filename}")

class ClusterData:
    def __init__(self, clustername: str, app_instance: AppClass, **kwargs):
        self.name = clustername
        self.cluster_type = ''
        for name, value in kwargs.items():
            setattr(self, name, value)
        self.fetched_data = {}
        self.app_instance = app_instance

    def gather_data(self):
        logging.info(f"Gathering data for {self.name}")
        user, enc = self.app_instance.config.get_user('clusters', self.name)
        try:
            with HostConnection(self.ip, username=user, password=enc, verify=False): # pyright: ignore[reportAttributeAccessIssue]

                self.fetched_data['locks'] = []
                locks = ClientLock.get_collection(**{"fields":"*"}) # pyright: ignore[reportArgumentType]
                for lock in locks:
                    self.fetched_data['locks'].append(lock.to_dict())
                logging.info(f"Locks found: {len(self.fetched_data['locks'])}")
        except Exception as e:
            logging.error(f"Could not retrieve data for {self.name} {e}", exc_info=e)

    def output_data(self):
        ws = self.app_instance.wb.create_sheet(self.name)
        ws.append(["Volume","Protocol","Type","Path","Lock","State","IP address"])
        for item in self.fetched_data['locks']:
            #'Volume,Protocol,Type,Path,Share Lock,Oplock Level,State,IP address'

            try:
                if item['type'] == 'share_level':
                    ws.append([item['volume']['name'],item['path'],item['protocol'],item['type'],f"{item['share_lock']}",item['state'],item['client_address']])
                else:
                    if item['type'] == 'op_lock':
                        ws.append([item['volume']['name'],item['path'],item['protocol'],item['type'],f"{item['oplock_level']}",item['state'],item['client_address']])
                    else:
                        logging.info(f"Unknown type for {item}")

            except Exception as e:
                logging.error(f"Lock error {e}")
                pprint.pprint(item)

if __name__ == '__main__':
    args = argp(script_name=script_name, description="get locks for a cluster(s)")
    config = Config(args.config_dir, args.output_dir) # pyright: ignore[reportAttributeAccessIssue]
    items = config.get_clusters(args.filter)
    # aiqums = config.search('aiqums', {'bu':'PUMA'})
    # pprint.pprint(aiqums)
    # connectors = config.search('connectors', {'bu':'PUMA'})
    # pprint.pprint(connectors)

    # pprint.pprint(items)

    APP = AppClass(script_name, items, config)
    APP.go()

